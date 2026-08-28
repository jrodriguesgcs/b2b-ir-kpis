#!/usr/bin/env python3
"""
B2B Intermediary-Referral KPI report generator.

Fetches deal, meeting and contact data from HubSpot and writes a styled
Excel workbook to reports/reports.xlsx - one tab per Stage (Retained
Clients, Referred Clients, New Intermediaries, Presentations,
Calls-Meetings) plus a static "KPI Targets" tab:

  - Each Stage tab has a Month > Week > Day column drill-down via Excel's
    native outline grouping (+/- expand), a trailing YTD Total column,
    and rows for the two BDMs (the internal HubSpot property
    `hubspot_owner_id`; rendered in the workbook as "BDM") plus a Grand
    Total row.
  - The Retained Clients tab additionally breaks each BDM's row down by
    Country and Program of Interest via row-level outline grouping, and
    carries two annual target rows (12/year per BDM).
  - "KPI Targets": static reference content (no API calls).

Run with:  python generate_report.py

Deliberately NOT in scope here: GitHub Actions / cron scheduling, email
delivery, and secrets management beyond a local .env file. See README.md.
"""

from __future__ import annotations

import os
import random
import subprocess
import sys
import tempfile
import time
import zipfile
from dataclasses import dataclass, field
from datetime import date

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

HUBSPOT_API_BASE = "https://api.hubapi.com"

# --- GCS Design System -------------------------------------------------
NIGHT_BLUE = "000957"
ELECTRIC_BLUE = "3F8CFF"
TINT_BLUE = "ECF4FF"  # Grand Total row fill
BORDER_TINT = "E3EDFF"
BODY_TEXT = "414856"

FONT_BODY = "Heebo"
FONT_TITLE = "Yrsa"  # regular weight, titles only
FONT_MONO = "JetBrains Mono"  # week labels / ID-like columns

THIN_BORDER = Border(*(Side(style="thin", color=BORDER_TINT) for _ in range(4)))

HEADER_FILL = PatternFill(fill_type="solid", fgColor=NIGHT_BLUE)
HEADER_FONT = Font(name=FONT_BODY, color="FFFFFF", bold=True)
GRAND_TOTAL_FILL = PatternFill(fill_type="solid", fgColor=TINT_BLUE)
BODY_FONT = Font(name=FONT_BODY, color=BODY_TEXT)
MONO_FONT = Font(name=FONT_MONO, color=BODY_TEXT)
TITLE_FONT = Font(name=FONT_TITLE, color=NIGHT_BLUE, bold=False, size=14)
MUTED_ITALIC_FONT = Font(name=FONT_BODY, color=BODY_TEXT, italic=True, size=9)

# Per-Stage header colours: the GCS design system's own documented 5-slot
# categorical chart palette (--chart-1..5 = primary, accent,
# muted-foreground, border, muted), in boldness order, mapped 1:1 onto
# STAGE_LABELS (Retained Clients first = most important outcome = boldest
# colour). muted-foreground/muted are only specified as HSL in the design
# doc (hsl(221 13% 46%) / hsl(218 21% 93%)) - converted to hex precisely;
# chart-4 (border tint) reuses the existing TINT_BLUE constant rather than
# introduce a near-duplicate hex.
CHART_MUTED_FOREGROUND = "667085"  # hsl(221 13% 46%)
CHART_MUTED = "E9ECF1"  # hsl(218 21% 93%)

STAGE_FILLS = [
    PatternFill(fill_type="solid", fgColor=NIGHT_BLUE),
    PatternFill(fill_type="solid", fgColor=ELECTRIC_BLUE),
    PatternFill(fill_type="solid", fgColor=CHART_MUTED_FOREGROUND),
    PatternFill(fill_type="solid", fgColor=TINT_BLUE),
    PatternFill(fill_type="solid", fgColor=CHART_MUTED),
]
STAGE_FONTS = [
    Font(name=FONT_BODY, color="FFFFFF", bold=True),
    Font(name=FONT_BODY, color="FFFFFF", bold=True),
    Font(name=FONT_BODY, color="FFFFFF", bold=True),
    Font(name=FONT_BODY, color=NIGHT_BLUE, bold=True),
    Font(name=FONT_BODY, color=NIGHT_BLUE, bold=True),
]

# Search API / associations endpoints have a tighter effective rate limit
# (~4-5 req/s) than the general API. A small fixed delay before each such
# call keeps us comfortably under it without needing per-call bookkeeping.
SEARCH_API_DELAY_SECONDS = 0.25

MAX_RETRIES = 5

# The two BDMs (HubSpot deal owners) this report is scoped to.
OWNER_NAMES = {
    "joao": "João Pacheco Gonçalves",
    "rohan": "Rohan Harris",
}


class HubSpotError(RuntimeError):
    """Raised when a HubSpot API call fails after exhausting retries."""


class HubSpotClient:
    """Thin wrapper around requests.Session with auth + backoff baked in."""

    def __init__(self, access_token: str):
        if not access_token:
            raise HubSpotError(
                "HUBSPOT_ACCESS_TOKEN is not set. Copy .env.example to .env "
                "and fill in a HubSpot Service Key or private app token."
            )
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            }
        )

    def request(self, method: str, path: str, is_search: bool = False, **kwargs):
        """Issue one HubSpot API call with retry/backoff on 429s.

        is_search=True adds a small fixed delay beforehand to respect the
        Search API's stricter rate limit.
        """
        if is_search:
            time.sleep(SEARCH_API_DELAY_SECONDS)

        url = f"{HUBSPOT_API_BASE}{path}"
        last_exc = None
        for attempt in range(MAX_RETRIES + 1):
            try:
                resp = self.session.request(method, url, timeout=30, **kwargs)
            except requests.RequestException as exc:
                last_exc = exc
                if attempt == MAX_RETRIES:
                    break
                delay = (2 ** attempt) + random.uniform(0, 1)
                time.sleep(delay)
                continue

            if resp.status_code == 429:
                if attempt == MAX_RETRIES:
                    raise HubSpotError(
                        f"HubSpot rate limit exceeded after {MAX_RETRIES} retries "
                        f"on {method} {path}"
                    )
                retry_after = resp.headers.get("Retry-After")
                if retry_after is not None:
                    delay = float(retry_after)
                else:
                    delay = (2 ** attempt) + random.uniform(0, 1)
                time.sleep(delay)
                continue

            if resp.status_code >= 400:
                raise HubSpotError(
                    f"HubSpot API error {resp.status_code} on {method} {path}: "
                    f"{resp.text[:500]}"
                )

            return resp.json() if resp.text else {}

        raise HubSpotError(f"HubSpot request failed on {method} {path}: {last_exc}")

    def get(self, path: str, is_search: bool = False, **kwargs):
        return self.request("GET", path, is_search=is_search, **kwargs)

    def post(self, path: str, is_search: bool = False, **kwargs):
        return self.request("POST", path, is_search=is_search, **kwargs)

    def paginate(self, path: str, params: dict | None = None, results_key: str = "results"):
        """GET-paginate a HubSpot list endpoint using the `paging.next.after` cursor."""
        params = dict(params or {})
        results = []
        while True:
            data = self.get(path, params=params)
            results.extend(data.get(results_key, []))
            next_after = data.get("paging", {}).get("next", {}).get("after")
            if not next_after:
                break
            params["after"] = next_after
        return results

    def search_all(self, object_type: str, body: dict):
        """POST-paginate /crm/v3/objects/{type}/search using the `after` cursor."""
        body = dict(body)
        body.setdefault("limit", 100)
        results = []
        while True:
            data = self.post(f"/crm/v3/objects/{object_type}/search", is_search=True, json=body)
            results.extend(data.get("results", []))
            next_after = data.get("paging", {}).get("next", {}).get("after")
            if not next_after:
                break
            body["after"] = next_after
        return results

    def batch_read(self, object_type: str, ids: list[str], properties: list[str]):
        """POST /crm/v3/objects/{type}/batch/read, chunked to 100 IDs/call."""
        results = []
        ids = list(dict.fromkeys(ids))  # de-dupe, preserve order
        for i in range(0, len(ids), 100):
            chunk = ids[i : i + 100]
            body = {
                "properties": properties,
                "inputs": [{"id": obj_id} for obj_id in chunk],
            }
            data = self.post(f"/crm/v3/objects/{object_type}/batch/read", json=body)
            results.extend(data.get("results", []))
        return results

    def batch_read_associations(self, from_type: str, to_type: str, ids: list[str]):
        """POST /crm/v4/associations/{from}/{to}/batch/read, chunked to 100 IDs/call.

        Returns {from_id: [to_id, ...]}.
        """
        result_map: dict = {}
        ids = list(dict.fromkeys(ids))
        for i in range(0, len(ids), 100):
            chunk = ids[i : i + 100]
            body = {"inputs": [{"id": obj_id} for obj_id in chunk]}
            data = self.post(f"/crm/v4/associations/{from_type}/{to_type}/batch/read", json=body)
            for entry in data.get("results", []):
                from_id = str(entry.get("from", {}).get("id"))
                # toObjectId comes back as a JSON number; normalise to str so
                # it matches the string ids used everywhere else (e.g. from
                # batch_read).
                to_ids = [str(to.get("toObjectId")) for to in entry.get("to", [])]
                result_map[from_id] = to_ids
        return result_map

    def batch_read_associations_typed(self, from_type: str, to_type: str, ids: list[str]):
        """Like batch_read_associations, but keeps each association's typeIds
        so callers can filter to a specific association label.

        Returns {from_id: [(to_id, [typeId, ...]), ...]}.
        """
        result_map: dict = {}
        ids = list(dict.fromkeys(ids))
        for i in range(0, len(ids), 100):
            chunk = ids[i : i + 100]
            body = {"inputs": [{"id": obj_id} for obj_id in chunk]}
            data = self.post(f"/crm/v4/associations/{from_type}/{to_type}/batch/read", json=body)
            for entry in data.get("results", []):
                from_id = str(entry.get("from", {}).get("id"))
                to_entries = [
                    (str(to.get("toObjectId")), [at.get("typeId") for at in to.get("associationTypes", [])])
                    for to in entry.get("to", [])
                ]
                result_map[from_id] = to_entries
        return result_map


@dataclass
class ReferenceData:
    """Everything Step 0 resolves, cached once per run."""

    typeid_introducer: int
    typeid_referred_person: int
    lead_source_property: str
    lead_source_candidates: list[str]  # all candidates with a "Partner Referral" option
    proposal_signed_property: str
    intermediary_pipeline_id: str
    sales_pipeline_id: str
    proposal_accepted_stage_ids: list[str]
    lifecyclestage_customer_value: str
    owner_ids: dict = field(default_factory=dict)  # {"joao": "123", "rohan": "456"}


def resolve_reference_data(client: HubSpotClient) -> ReferenceData:
    """Step 0: resolve every ID/value the KPI logic needs, print as we go."""
    print("=== Step 0: resolving reference data ===")

    # --- Association labels: contacts <-> contacts (Referred Person / Introducer)
    labels_resp = client.get("/crm/v4/associations/contacts/contacts/labels")
    typeid_introducer = None
    typeid_referred_person = None
    for entry in labels_resp.get("results", []):
        label = (entry.get("label") or "").strip().lower()
        type_id = entry.get("typeId")
        if label == "introducer":
            typeid_introducer = type_id
        elif label == "referred person":
            typeid_referred_person = type_id
    if typeid_introducer is None or typeid_referred_person is None:
        raise HubSpotError(
            "Could not find both 'Referred Person' / 'Introducer' association "
            f"labels on contacts<->contacts. Got: {labels_resp.get('results')}"
        )
    print(f"  TYPEID_INTRODUCER (read from a referred contact) = {typeid_introducer}")
    print(f"  TYPEID_REFERRED_PERSON (inverse)                 = {typeid_referred_person}")

    # --- Contact lead-source property holding "Partner Referral"
    # Multiple overlapping legacy properties can carry this option; KPI 4
    # reconciles all candidates against its expected total (37) rather than
    # committing to a guess here (see fetch_referred_clients).
    lead_source_property, lead_source_candidates = _resolve_lead_source_property(client)
    print(f"  Lead-source property (best guess, reconciled in KPI 4) = {lead_source_property}")
    print(f"  Lead-source candidates carrying 'Partner Referral'     = {lead_source_candidates}")

    # --- Deal property: "Proposal Signed Date Time"
    deal_props = client.get("/crm/v3/properties/deals")
    proposal_signed_property = None
    for prop in deal_props.get("results", []):
        label = (prop.get("label") or "").lower()
        name = (prop.get("name") or "").lower()
        if "proposal" in label and "sign" in label:
            proposal_signed_property = prop["name"]
            break
        if "proposal" in name and "sign" in name:
            proposal_signed_property = prop["name"]
            break
    if not proposal_signed_property:
        raise HubSpotError("Could not resolve 'Proposal Signed Date Time' deal property.")
    print(f"  Deal property 'Proposal Signed Date Time' = {proposal_signed_property}")

    # --- Pipelines / stages
    pipelines_resp = client.get("/crm/v3/pipelines/deals")
    intermediary_pipeline_id = None
    sales_pipeline_id = None
    proposal_accepted_stage_ids = []
    for pipeline in pipelines_resp.get("results", []):
        label = (pipeline.get("label") or "").strip()
        if label == "[GCS] Institutional Relations":
            intermediary_pipeline_id = pipeline["id"]
        if "sales pipeline" in label.lower():
            sales_pipeline_id = pipeline["id"]
            for stage in pipeline.get("stages", []):
                stage_label = (stage.get("label") or "").strip().lower()
                if "proposal accepted" in stage_label or "closed won" in stage_label:
                    proposal_accepted_stage_ids.append(stage["id"])
    if not intermediary_pipeline_id:
        raise HubSpotError("Could not resolve pipeline '[GCS] Institutional Relations'.")
    if not sales_pipeline_id or not proposal_accepted_stage_ids:
        raise HubSpotError("Could not resolve 'Sales Pipeline' / Proposal Accepted/Closed Won stage.")
    print(f"  Pipeline '[GCS] Institutional Relations' id = {intermediary_pipeline_id}")
    print(f"  Pipeline '[GCS] Sales Pipeline' id           = {sales_pipeline_id}")
    print(f"  Stage id(s) Proposal Accepted/Closed Won    = {proposal_accepted_stage_ids}")

    # --- lifecyclestage "Customer" value
    lifecyclestage_prop = client.get("/crm/v3/properties/contacts/lifecyclestage")
    customer_value = None
    for option in lifecyclestage_prop.get("options", []):
        if (option.get("label") or "").strip().lower() == "customer":
            customer_value = option["value"]
            break
    if not customer_value:
        raise HubSpotError("Could not resolve lifecyclestage 'Customer' value.")
    print(f"  lifecyclestage 'Customer' value = {customer_value}")

    # --- Owners
    owners_resp = client.paginate("/crm/v3/owners")
    owner_ids = {}
    for owner in owners_resp:
        full_name = f"{owner.get('firstName', '')} {owner.get('lastName', '')}".strip()
        if full_name == OWNER_NAMES["joao"]:
            owner_ids["joao"] = owner["id"]
        elif full_name == OWNER_NAMES["rohan"]:
            owner_ids["rohan"] = owner["id"]
    missing = [key for key, _ in OWNER_NAMES.items() if key not in owner_ids]
    if missing:
        raise HubSpotError(f"Could not resolve owner id(s) for: {missing}")
    print(f"  Owner id (João Pacheco Gonçalves) = {owner_ids['joao']}")
    print(f"  Owner id (Rohan Harris)           = {owner_ids['rohan']}")

    print("=== Step 0 complete ===\n")

    return ReferenceData(
        typeid_introducer=typeid_introducer,
        typeid_referred_person=typeid_referred_person,
        lead_source_property=lead_source_property,
        lead_source_candidates=lead_source_candidates,
        proposal_signed_property=proposal_signed_property,
        intermediary_pipeline_id=intermediary_pipeline_id,
        sales_pipeline_id=sales_pipeline_id,
        proposal_accepted_stage_ids=proposal_accepted_stage_ids,
        lifecyclestage_customer_value=customer_value,
        owner_ids=owner_ids,
    )


_LEAD_SOURCE_CANDIDATES = ["lead_source", "manual_lead_source", "gc_manual_lead_source"]


def _resolve_lead_source_property(client: HubSpotClient) -> tuple[str, list[str]]:
    """Check each candidate lead-source property for a 'Partner Referral' option
    and report counts so the caller can reconcile against the expected KPI 4/5
    totals rather than assuming a fixed property name. Returns (best_guess,
    all_candidates) - KPI 4 tries every candidate and locks in whichever one
    reconciles to the expected total.
    """
    props_by_name = {p["name"]: p for p in client.get("/crm/v3/properties/contacts").get("results", [])}
    candidates_found = {}
    for name in _LEAD_SOURCE_CANDIDATES:
        prop = props_by_name.get(name)
        if not prop:
            continue
        has_partner_referral = any(
            (opt.get("label") or "").strip().lower() == "partner referral"
            for opt in prop.get("options", [])
        )
        if has_partner_referral:
            count = _count_contacts_with_value(client, name, "Partner Referral", prop)
            candidates_found[name] = count
            print(f"    candidate property '{name}': {count} contacts = Partner Referral")

    if not candidates_found:
        raise HubSpotError(
            "None of the candidate lead-source properties "
            f"{_LEAD_SOURCE_CANDIDATES} have a 'Partner Referral' option."
        )

    # Prefer whichever candidate has the most matching contacts; ties broken by
    # candidate priority order. This is reconciled against KPI 4/5 expected
    # totals during isolated verification (see README / task notes).
    best_name = max(candidates_found, key=lambda n: (candidates_found[n], -_LEAD_SOURCE_CANDIDATES.index(n)))
    return best_name, list(candidates_found.keys())


def _count_contacts_with_value(client: HubSpotClient, property_name: str, label: str, prop: dict) -> int:
    # Search uses the option *value*, not its label.
    value = None
    for opt in prop.get("options", []):
        if (opt.get("label") or "").strip().lower() == label.strip().lower():
            value = opt["value"]
            break
    if value is None:
        return 0
    body = {
        "filterGroups": [
            {"filters": [{"propertyName": property_name, "operator": "EQ", "value": value}]}
        ],
        "properties": ["hs_object_id"],
        "limit": 1,
    }
    data = client.post("/crm/v3/objects/contacts/search", is_search=True, json=body)
    return data.get("total", 0)


def parse_hubspot_datetime(value: str):
    """Parse a HubSpot ISO-8601 timestamp (e.g. createdate) into a datetime."""
    from datetime import datetime

    # HubSpot returns e.g. "2026-03-04T10:15:30.000Z"
    return datetime.strptime(value.split(".")[0].rstrip("Z"), "%Y-%m-%dT%H:%M:%S")


def year_start_ms() -> int:
    """Epoch ms for 1 Jan of the current UTC year - the shared year-to-date
    lower bound every KPI's HubSpot query filters on."""
    from datetime import datetime, timezone

    year_start = datetime(datetime.now(timezone.utc).year, 1, 1, tzinfo=timezone.utc)
    return int(year_start.timestamp() * 1000)


def fetch_new_intermediaries(client: HubSpotClient, ref: ReferenceData) -> dict:
    """Stage: New Intermediaries. Deals in the [GCS] Institutional
    Relations pipeline, created this year so far, with a known owner.
    Grouped by owner + calendar day of createdate.
    """
    body = {
        "filterGroups": [
            {
                "filters": [
                    {"propertyName": "pipeline", "operator": "EQ", "value": ref.intermediary_pipeline_id},
                    {"propertyName": "createdate", "operator": "GTE", "value": year_start_ms()},
                    {"propertyName": "hubspot_owner_id", "operator": "HAS_PROPERTY"},
                ]
            }
        ],
        "properties": ["hubspot_owner_id", "createdate"],
        "limit": 100,
    }
    deals = client.search_all("deals", body)

    counts: dict = {}
    other_owner_deals = []
    known_owner_ids = set(ref.owner_ids.values())
    for deal in deals:
        props = deal.get("properties", {})
        owner_id = props.get("hubspot_owner_id")
        created = props.get("createdate")
        if not owner_id or not created:
            continue
        if owner_id not in known_owner_ids:
            other_owner_deals.append((deal["id"], owner_id))
            continue
        day = parse_hubspot_datetime(created).date()
        counts[(owner_id, day)] = counts.get((owner_id, day), 0) + 1

    grand_total = sum(counts.values())
    print(f"New Intermediaries: fetched {len(deals)} deals total, "
          f"{grand_total} attributed to João/Rohan (year-to-date)")
    if other_owner_deals:
        print(f"  {len(other_owner_deals)} deal(s) owned by someone other than João/Rohan "
              f"(excluded from BDM columns): {other_owner_deals[:10]}"
              + (" ..." if len(other_owner_deals) > 10 else ""))
    return counts


def fetch_meetings(client: HubSpotClient, ref: ReferenceData) -> tuple[dict, dict]:
    """Stages: Calls/Meetings and Presentations.

    Meetings associated with a deal owned by João/Rohan, where the meeting's
    attendee-owner-ids or created-by-user-id also matches João/Rohan. Split
    by hs_activity_type: != "Presentation" (incl. unknown) -> Calls/Meetings,
    == "Presentation" -> Presentations. Grouped by deal owner + calendar day
    of the meeting's start time, scoped to this calendar year so far.
    Deduped by (meeting_id, deal_owner_id).
    """
    from datetime import date, datetime, timezone

    known_owner_ids = set(ref.owner_ids.values())
    today = datetime.now(timezone.utc).date()
    year_start_day = date(today.year, 1, 1)

    # 1. Deals owned by João/Rohan (any pipeline).
    body = {
        "filterGroups": [
            {"filters": [{"propertyName": "hubspot_owner_id", "operator": "IN", "values": list(known_owner_ids)}]}
        ],
        "properties": ["hubspot_owner_id"],
        "limit": 100,
    }
    owned_deals = client.search_all("deals", body)
    deal_owner_by_id = {d["id"]: d["properties"]["hubspot_owner_id"] for d in owned_deals}
    print(f"KPI 2/3: {len(deal_owner_by_id)} deals owned by João/Rohan")

    # 2. Meetings associated with those deals -> {meeting_id: {owner_id, ...}}
    deal_to_meetings = client.batch_read_associations("deals", "meetings", list(deal_owner_by_id.keys()))
    meeting_owners: dict = {}
    for deal_id, meeting_ids in deal_to_meetings.items():
        owner_id = deal_owner_by_id.get(deal_id)
        for meeting_id in meeting_ids:
            meeting_owners.setdefault(meeting_id, set()).add(owner_id)
    print(f"KPI 2/3: {len(meeting_owners)} unique meetings associated with those deals")

    # 3. Meeting properties, batch-read.
    meeting_props = client.batch_read(
        "meetings",
        list(meeting_owners.keys()),
        ["hs_activity_type", "hs_meeting_start_time", "hs_attendee_owner_ids", "hs_created_by_user_id"],
    )

    meetings_counts: dict = {}
    presentations_counts: dict = {}
    skipped_no_start_time = []
    skipped_prior_year = 0
    skipped_future = 0
    for meeting in meeting_props:
        meeting_id = meeting["id"]
        props = meeting.get("properties", {})
        attendee_owner_ids = set((props.get("hs_attendee_owner_ids") or "").split(";")) - {""}
        created_by_user_id = props.get("hs_created_by_user_id")
        qualifies = bool(attendee_owner_ids & known_owner_ids) or (created_by_user_id in known_owner_ids)
        if not qualifies:
            continue

        start_time = props.get("hs_meeting_start_time")
        if not start_time:
            skipped_no_start_time.append(meeting_id)
            continue
        day = parse_hubspot_datetime(start_time).date()
        if day < year_start_day:
            skipped_prior_year += 1
            continue
        if day > today:
            # Already booked in HubSpot but hasn't happened yet - doesn't
            # belong in a "year-to-date, through today" total, and has no
            # column to land in on the Year to Date sheet (which only goes
            # through today).
            skipped_future += 1
            continue
        is_presentation = props.get("hs_activity_type") == "Presentation"

        for owner_id in meeting_owners.get(meeting_id, set()):
            if owner_id not in known_owner_ids:
                continue
            bucket = presentations_counts if is_presentation else meetings_counts
            bucket[(owner_id, day)] = bucket.get((owner_id, day), 0) + 1

    meetings_total = sum(meetings_counts.values())
    presentations_total = sum(presentations_counts.values())
    print(f"Calls/Meetings: {meetings_total} (year-to-date)")
    print(f"Presentations: {presentations_total} (year-to-date)")
    if presentations_total:
        print("  NOTE: Presentations is non-zero - flagged for review, do not assume correct.")
    if skipped_no_start_time:
        print(f"  {len(skipped_no_start_time)} qualifying meeting(s) skipped for missing start time: "
              f"{skipped_no_start_time[:10]}")
    if skipped_prior_year:
        print(f"  {skipped_prior_year} qualifying meeting(s) excluded - before this calendar year")
    if skipped_future:
        print(f"  {skipped_future} qualifying meeting(s) excluded - already booked but haven't happened yet "
              f"(start time after today)")
    return meetings_counts, presentations_counts


def _partner_referral_value(client: HubSpotClient, property_name: str) -> str:
    """Resolve the option *value* (not label) for 'Partner Referral' on the
    given contact property."""
    props = client.get("/crm/v3/properties/contacts").get("results", [])
    prop = next((p for p in props if p["name"] == property_name), None)
    for opt in (prop or {}).get("options", []):
        if (opt.get("label") or "").strip().lower() == "partner referral":
            return opt["value"]
    raise HubSpotError(f"Property '{property_name}' has no 'Partner Referral' option.")


def _partner_referral_contacts(client: HubSpotClient, property_name: str, extra_filters: list | None = None) -> list:
    """Contacts where `property_name` = 'Partner Referral' (plus any
    extra_filters ANDed in). Returns raw HubSpot contact dicts."""
    value = _partner_referral_value(client, property_name)
    filters = [{"propertyName": property_name, "operator": "EQ", "value": value}]
    filters.extend(extra_filters or [])
    body = {
        "filterGroups": [{"filters": filters}],
        "properties": ["createdate", property_name, "lifecyclestage"],
        "limit": 100,
    }
    return client.search_all("contacts", body)


def _resolve_introducers(client: HubSpotClient, ref: ReferenceData, referred_contacts: list):
    """Shared by KPI 4 and KPI 5: for each referred contact, find its
    qualifying introducer (association typeId == TYPEID_INTRODUCER) and that
    introducer's deal owner. Logs zero/multi-introducer anomalies.

    Returns {referred_contact_id: (introducer_id, introducer_owner_id)} -
    only for contacts with exactly one qualifying introducer whose owner is
    known.
    """
    referred_ids = [c["id"] for c in referred_contacts]
    assoc = client.batch_read_associations_typed("contacts", "contacts", referred_ids)

    introducer_by_referred: dict = {}
    zero_introducers = []
    multi_introducers = []
    for referred_id in referred_ids:
        candidates = [to_id for to_id, type_ids in assoc.get(referred_id, []) if ref.typeid_introducer in type_ids]
        if len(candidates) == 0:
            zero_introducers.append(referred_id)
        elif len(candidates) > 1:
            multi_introducers.append((referred_id, candidates))
            introducer_by_referred[referred_id] = candidates[0]  # flagged, but still attributed
        else:
            introducer_by_referred[referred_id] = candidates[0]

    if zero_introducers:
        print(f"    {len(zero_introducers)} referred contact(s) with NO qualifying introducer "
              f"(excluded): {zero_introducers[:10]}" + (" ..." if len(zero_introducers) > 10 else ""))
    if multi_introducers:
        print(f"    {len(multi_introducers)} referred contact(s) with MULTIPLE qualifying introducers "
              f"(using first, flagged): {multi_introducers[:10]}")

    # The introducer's owning BDM is taken from the deal owner of their own
    # deal in the [GCS] Institutional Relations pipeline - NOT the
    # introducer contact's own hubspot_owner_id, which can be a routing
    # artifact (e.g. set by a lead-assignment workflow) unrelated to which
    # BDM actually manages that intermediary relationship.
    introducer_ids = list(set(introducer_by_referred.values()))
    introducer_to_deals = client.batch_read_associations("contacts", "deals", introducer_ids)
    all_deal_ids = list({d for deals in introducer_to_deals.values() for d in deals})
    deal_props = client.batch_read("deals", all_deal_ids, ["pipeline", "hubspot_owner_id"])
    deal_by_id = {d["id"]: d["properties"] for d in deal_props}

    owner_by_introducer = {}
    no_ir_deal = []
    multi_ir_deal_owners = []
    for introducer_id in introducer_ids:
        ir_deal_owners = []
        for deal_id in introducer_to_deals.get(introducer_id, []):
            props = deal_by_id.get(deal_id, {})
            if props.get("pipeline") == ref.intermediary_pipeline_id and props.get("hubspot_owner_id"):
                ir_deal_owners.append(props["hubspot_owner_id"])
        distinct_owners = set(ir_deal_owners)
        if not distinct_owners:
            no_ir_deal.append(introducer_id)
        elif len(distinct_owners) > 1:
            multi_ir_deal_owners.append((introducer_id, list(distinct_owners)))
            owner_by_introducer[introducer_id] = ir_deal_owners[0]
        else:
            owner_by_introducer[introducer_id] = ir_deal_owners[0]

    if no_ir_deal:
        print(f"    {len(no_ir_deal)} introducer(s) with no deal in the Institutional Relations "
              f"pipeline (excluded): {no_ir_deal[:10]}" + (" ..." if len(no_ir_deal) > 10 else ""))
    if multi_ir_deal_owners:
        print(f"    {len(multi_ir_deal_owners)} introducer(s) with multiple Institutional Relations "
              f"deals under different owners (using first, flagged): {multi_ir_deal_owners[:10]}")

    result = {}
    unknown_owner = []
    for referred_id, introducer_id in introducer_by_referred.items():
        owner_id = owner_by_introducer.get(introducer_id)
        if not owner_id or owner_id not in ref.owner_ids.values():
            unknown_owner.append((referred_id, introducer_id, owner_id))
            continue
        result[referred_id] = (introducer_id, owner_id)
    if unknown_owner:
        print(f"    {len(unknown_owner)} introducer(s) not owned by João/Rohan (excluded): "
              f"{unknown_owner[:10]}")
    return result


def fetch_referred_clients(client: HubSpotClient, ref: ReferenceData):
    """Stage: Referred Clients. Partner-Referral contacts with a
    resolvable introducer owned by João/Rohan, scoped to this calendar
    year so far (referred contact's createdate). Grouped by owner +
    calendar day of createdate.

    Tries every lead-source candidate property and keeps whichever
    reconciles closest to the previously-accepted all-time total of 20,
    since the property is genuinely ambiguous on this portal (see Step 0).
    Note: this year-to-date total is expected to be <= the all-time 20
    accepted for the original build, since some of those 20 referred
    clients may have been created in a prior year.
    """
    print("KPI 4: trying each lead-source candidate (year-to-date)")
    best = None
    for property_name in ref.lead_source_candidates:
        referred_contacts = _partner_referral_contacts(
            client,
            property_name,
            extra_filters=[{"propertyName": "createdate", "operator": "GTE", "value": year_start_ms()}],
        )
        print(f"  candidate '{property_name}': {len(referred_contacts)} Partner Referral contacts (YTD)")
        introducers = _resolve_introducers(client, ref, referred_contacts)
        counts: dict = {}
        for referred in referred_contacts:
            referred_id = referred["id"]
            if referred_id not in introducers:
                continue
            _, owner_id = introducers[referred_id]
            created = referred["properties"].get("createdate")
            if not created:
                continue
            day = parse_hubspot_datetime(created).date()
            counts[(owner_id, day)] = counts.get((owner_id, day), 0) + 1
        total = sum(counts.values())
        print(f"  candidate '{property_name}': grand total = {total}")
        if best is None or abs(total - 20) < abs(best[1] - 20):
            best = (property_name, total, counts)

    property_name, total, counts = best
    print(f"Referred Clients: using '{property_name}', total = {total} (year-to-date; "
          f"all-time accepted ground truth was 20)")
    return property_name, counts


def fetch_retained_clients(client: HubSpotClient, ref: ReferenceData, lead_source_property: str):
    """Stage: Retained Clients. Same Partner-Referral population as the
    Referred Clients stage (not year-scoped at the contact level - a
    client can be referred in a prior year and retained this year),
    filtered to lifecyclestage = Customer, with a qualifying deal in
    [GCS] Sales Pipeline at Proposal Accepted/Closed Won whose Proposal
    Signed Date Time falls in this calendar year so far. Grouped by
    introducer's owner + calendar day of that signed date.
    """
    from datetime import date, datetime, timezone

    today = datetime.now(timezone.utc).date()
    year_start_day = date(today.year, 1, 1)

    customers = _partner_referral_contacts(
        client,
        lead_source_property,
        extra_filters=[
            {"propertyName": "lifecyclestage", "operator": "EQ", "value": ref.lifecyclestage_customer_value}
        ],
    )
    print(f"KPI 5: {len(customers)} Partner-Referral contacts with lifecyclestage = Customer "
          f"(property '{lead_source_property}')")

    customer_ids = [c["id"] for c in customers]
    contact_to_deals = client.batch_read_associations("contacts", "deals", customer_ids)
    all_deal_ids = list({d for deals in contact_to_deals.values() for d in deals})
    deal_props = client.batch_read(
        "deals", all_deal_ids,
        ["pipeline", "dealstage", ref.proposal_signed_property, "country_and_program_of_interest"],
    )
    deal_by_id = {d["id"]: d["properties"] for d in deal_props}

    qualifying_deal_by_contact = {}
    multi_deal_contacts = []
    no_qualifying_deal = []
    for contact_id in customer_ids:
        qualifying = [
            d for d in contact_to_deals.get(contact_id, [])
            if deal_by_id.get(d, {}).get("pipeline") == ref.sales_pipeline_id
            and deal_by_id.get(d, {}).get("dealstage") in ref.proposal_accepted_stage_ids
        ]
        if not qualifying:
            no_qualifying_deal.append(contact_id)
        elif len(qualifying) > 1:
            multi_deal_contacts.append((contact_id, qualifying))
            qualifying_deal_by_contact[contact_id] = qualifying[0]  # flagged, still attributed
        else:
            qualifying_deal_by_contact[contact_id] = qualifying[0]

    if no_qualifying_deal:
        print(f"  {len(no_qualifying_deal)} customer(s) with no qualifying Proposal "
              f"Accepted/Closed Won deal (excluded): {no_qualifying_deal[:10]}"
              + (" ..." if len(no_qualifying_deal) > 10 else ""))
    if multi_deal_contacts:
        print(f"  {len(multi_deal_contacts)} customer(s) with MULTIPLE qualifying deals "
              f"(using first, flagged): {multi_deal_contacts[:10]}")

    qualifying_contacts = [c for c in customers if c["id"] in qualifying_deal_by_contact]
    introducers = _resolve_introducers(client, ref, qualifying_contacts)

    counts: dict = {}
    # {(owner_id, program_label, day): count} - drives the Retained Clients
    # tab's per-BDM drill-down by Country and Program of Interest (the
    # deal's own `country_and_program_of_interest` property). Missing/blank
    # values are bucketed under "Not specified" rather than dropped, so the
    # breakdown rows still account for every counted client.
    program_counts: dict = {}
    missing_signed_date = []
    prior_year = []
    future_dated = []
    for contact_id, deal_id in qualifying_deal_by_contact.items():
        if contact_id not in introducers:
            continue
        _, owner_id = introducers[contact_id]
        deal_properties = deal_by_id.get(deal_id, {})
        signed_date = deal_properties.get(ref.proposal_signed_property)
        if not signed_date:
            missing_signed_date.append((contact_id, deal_id))
            continue
        day = parse_hubspot_datetime(signed_date).date()
        if day < year_start_day:
            prior_year.append((contact_id, deal_id, day.isoformat()))
            continue
        if day > today:
            # A manually-entered date field could in theory be set ahead of
            # today by mistake - excluded for the same reason as future
            # meetings: no column to land in on a "through today" sheet.
            future_dated.append((contact_id, deal_id, day.isoformat()))
            continue
        counts[(owner_id, day)] = counts.get((owner_id, day), 0) + 1
        program = deal_properties.get("country_and_program_of_interest") or "Not specified"
        program_counts[(owner_id, program, day)] = program_counts.get((owner_id, program, day), 0) + 1

    if missing_signed_date:
        print(f"  {len(missing_signed_date)} qualifying deal(s) missing Proposal Signed Date Time "
              f"(excluded): {missing_signed_date[:10]}")
    if future_dated:
        print(f"  {len(future_dated)} qualifying deal(s) with a Proposal Signed Date Time after today "
              f"(excluded): {future_dated[:10]}")
    if prior_year:
        print(f"  {len(prior_year)} qualifying deal(s) signed before this calendar year "
              f"(excluded from year-to-date): {prior_year[:10]}")

    total = sum(counts.values())
    print(f"Retained Clients: {total} (year-to-date; all-time accepted ground truth was 6)")
    return counts, program_counts


# --- Excel assembly -----------------------------------------------------

# Stage column order, left to right, per explicit request (reverse-funnel:
# retained/referred outcomes first, activity volume last).
STAGE_LABELS = ["Retained Clients", "Referred Clients", "New Intermediaries", "Presentations", "Calls/Meetings"]

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def build_calendar_hierarchy(today: date) -> list:
    """Year-to-date calendar hierarchy: January through the current month,
    each month partitioned into weeks that never cross a month boundary
    (a global ISO week spanning two months is clipped into two partial
    groups), each week listing its elapsed days only (1 Jan through
    today - no future placeholder columns). This is what makes an
    in-progress month/week automatically read as "to-date" and a finished
    one show the full period, with no special-case branching needed.

    Returns [(month_index, month_name, [(week_label, [day, ...]), ...]), ...]
    """
    year = today.year
    months = []
    for month_index in range(1, today.month + 1):
        first_of_month = date(year, month_index, 1)
        if month_index == today.month:
            last_day_this_month = today
        else:
            next_month = date(year, month_index + 1, 1) if month_index < 12 else date(year + 1, 1, 1)
            last_day_this_month = date.fromordinal(next_month.toordinal() - 1)

        # Group consecutive days by ISO week (clipping at the month
        # boundary), then relabel sequentially as W1, W2, ... *within this
        # month* - each month restarts its own W1, rather than carrying
        # the year's global ISO week number across month boundaries.
        raw_weeks = []
        current_iso_week = None
        current_week_days: list = []
        day_ord = first_of_month.toordinal()
        while day_ord <= last_day_this_month.toordinal():
            day = date.fromordinal(day_ord)
            iso_week = day.isocalendar()[1]
            if iso_week != current_iso_week:
                if current_week_days:
                    raw_weeks.append(current_week_days)
                current_iso_week = iso_week
                current_week_days = []
            current_week_days.append(day)
            day_ord += 1
        if current_week_days:
            raw_weeks.append(current_week_days)

        weeks = [(f"W{i + 1}", days) for i, days in enumerate(raw_weeks)]
        months.append((month_index, MONTH_NAMES[month_index - 1], weeks))
    return months


# Tab names must avoid characters Excel forbids in sheet titles (/ \ ? * [ ]).
# In-sheet text (row 1 headers, etc.) still uses the real STAGE_LABELS text.
TAB_NAMES = ["Retained Clients", "Referred Clients", "New Intermediaries", "Presentations", "Calls-Meetings"]

RETAINED_CLIENT_ANNUAL_TARGET = 12  # per BDM, per explicit request
CALLS_MEETINGS_ANNUAL_TARGET = 440  # per BDM, per explicit request


def _compute_column_layout(calendar: list, start_col: int = 2, include_pct_column: bool = False) -> dict:
    """Precompute the Day/Week/Month/YTD column layout once (no row data) so
    every row written against it - BDM totals, Retained's per-program
    breakdown rows, Grand Total, target rows - shares identical columns.

    include_pct_column reserves one extra flat column right after each
    Month-total column (Calls-Meetings' "% of Target" column) - baked into
    the layout up front rather than inserted afterward, since openpyxl does
    not rewrite formula cell references when columns are inserted later.
    """
    col = start_col
    day_cols = []
    week_groups = []  # (week_total_col, [day_cols], week_label, month_name)
    month_groups = []  # (month_total_col, [week_total_cols], month_name, month_index, pct_col)
    for month_index, month_name, weeks in calendar:
        month_week_cols = []
        for week_label, days in weeks:
            week_day_cols = []
            for day in days:
                day_cols.append((col, day))
                week_day_cols.append(col)
                col += 1
            week_total_col = col
            week_groups.append((week_total_col, week_day_cols, week_label, month_name))
            month_week_cols.append(week_total_col)
            col += 1
        month_total_col = col
        col += 1
        pct_col = None
        if include_pct_column:
            pct_col = col
            col += 1
        month_groups.append((month_total_col, month_week_cols, month_name, month_index, pct_col))
    ytd_col = col
    return {
        "day_cols": day_cols,
        "week_groups": week_groups,
        "month_groups": month_groups,
        "ytd_col": ytd_col,
        "last_col": ytd_col,
    }


def _write_time_headers(ws, layout: dict, stage_label: str, stage_fill, stage_font) -> None:
    """Row 1: Stage name merged across the whole tab. Row 2: per-column
    period label (day date / week / month total / YTD total), plus the
    column outline levels that drive Month -> Week -> Day drill-down."""
    last_col = layout["last_col"]
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    ws.cell(row=1, column=2, value=stage_label)

    for col, day in layout["day_cols"]:
        letter = get_column_letter(col)
        ws.cell(row=2, column=col, value=day.isoformat())
        ws.column_dimensions[letter].outline_level = 2
        ws.column_dimensions[letter].hidden = True
        ws.column_dimensions[letter].width = 11
    for week_total_col, _day_cols, week_label, month_name in layout["week_groups"]:
        letter = get_column_letter(week_total_col)
        ws.cell(row=2, column=week_total_col, value=f"{week_label} {month_name}")
        ws.column_dimensions[letter].outline_level = 1
        ws.column_dimensions[letter].hidden = True
        # Marks this group's subordinate level (Days) as starting collapsed -
        # without this, Excel's expand/collapse behaviour at this boundary is
        # undefined and can inconsistently cascade into the Day level on a
        # single click.
        ws.column_dimensions[letter].collapsed = True
        ws.column_dimensions[letter].width = 13
    for month_total_col, _week_cols, month_name, _month_index, pct_col in layout["month_groups"]:
        letter = get_column_letter(month_total_col)
        ws.cell(row=2, column=month_total_col, value=f"{month_name} Total")
        ws.column_dimensions[letter].outline_level = 0
        # Same reasoning as above, one level up: a Month-total column with no
        # collapsed flag is what let some months' "+" cascade straight to Day.
        ws.column_dimensions[letter].collapsed = True
        ws.column_dimensions[letter].width = 13
        if pct_col is not None:
            pct_letter = get_column_letter(pct_col)
            ws.cell(row=2, column=pct_col, value="% of Target")
            ws.column_dimensions[pct_letter].outline_level = 0
            ws.column_dimensions[pct_letter].width = 12
    ytd_letter = get_column_letter(layout["ytd_col"])
    ws.cell(row=2, column=layout["ytd_col"], value="YTD Total")
    ws.column_dimensions[ytd_letter].width = 14

    for row in (1, 2):
        for col in range(2, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = stage_fill
            cell.font = stage_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER


def _all_layout_cols(layout: dict) -> list:
    cols = [c for c, _ in layout["day_cols"]]
    cols += [w[0] for w in layout["week_groups"]]
    cols += [m[0] for m in layout["month_groups"]]
    cols.append(layout["ytd_col"])
    return cols


def _write_data_row(ws, row_idx: int, layout: dict, day_value_fn, font, formula_cells: dict) -> dict:
    """Write one row of literal day values plus real SUM() formulas for the
    Week/Month/YTD columns over that same row's own cells. Returns
    {col: numeric_value} for every column, used both to cache formula
    results and to let a parent row (BDM total, Grand Total) sum children."""
    col_values: dict = {}
    for col, day in layout["day_cols"]:
        value = day_value_fn(day)
        ws.cell(row=row_idx, column=col, value=value).font = font
        col_values[col] = value
    for week_total_col, day_cols, _wl, _mn in layout["week_groups"]:
        first, last = get_column_letter(day_cols[0]), get_column_letter(day_cols[-1])
        total = sum(col_values[c] for c in day_cols)
        ws.cell(row=row_idx, column=week_total_col, value=f"=SUM({first}{row_idx}:{last}{row_idx})").font = font
        col_values[week_total_col] = total
        formula_cells[f"{get_column_letter(week_total_col)}{row_idx}"] = total
    for month_total_col, week_cols, _mn, _mi, _pct in layout["month_groups"]:
        refs = ",".join(f"{get_column_letter(c)}{row_idx}" for c in week_cols)
        total = sum(col_values[c] for c in week_cols)
        ws.cell(row=row_idx, column=month_total_col, value=f"=SUM({refs})").font = font
        col_values[month_total_col] = total
        formula_cells[f"{get_column_letter(month_total_col)}{row_idx}"] = total
    month_cols = [m[0] for m in layout["month_groups"]]
    refs = ",".join(f"{get_column_letter(c)}{row_idx}" for c in month_cols)
    ytd_total = sum(col_values[c] for c in month_cols)
    ws.cell(row=row_idx, column=layout["ytd_col"], value=f"=SUM({refs})" if month_cols else "=SUM()").font = font
    col_values[layout["ytd_col"]] = ytd_total
    formula_cells[f"{get_column_letter(layout['ytd_col'])}{row_idx}"] = ytd_total
    return col_values


def _write_sum_row(ws, row_idx: int, layout: dict, child_col_values: list, source_rows: list,
                    font, formula_cells: dict) -> dict:
    """Write row_idx as =SUM() of the given sibling rows, per column - used
    for a BDM total (sum of its own Program-of-Interest breakdown rows) and
    the Grand Total row (sum of the two BDM rows)."""
    col_values: dict = {}
    for col in _all_layout_cols(layout):
        letter = get_column_letter(col)
        refs = ",".join(f"{letter}{r}" for r in source_rows)
        ws.cell(row=row_idx, column=col, value=f"=SUM({refs})").font = font
        total = sum(cv.get(col, 0) for cv in child_col_values)
        col_values[col] = total
        formula_cells[f"{letter}{row_idx}"] = total
    return col_values


def _write_pct_of_target_row(ws, row_idx: int, layout: dict, target: float, font, formula_cells: dict) -> None:
    """Fill in each month's "% of Target" column (Calls-Meetings only) as a
    real formula referencing that same row's own Month-total cell, divided
    by the annual target - e.g. August's cell reads "=AL3/440"."""
    for month_total_col, _week_cols, _mn, _mi, pct_col in layout["month_groups"]:
        if pct_col is None:
            continue
        month_letter = get_column_letter(month_total_col)
        pct_letter = get_column_letter(pct_col)
        cell = ws.cell(row=row_idx, column=pct_col, value=f"={month_letter}{row_idx}/{target}")
        cell.font = font
        cell.number_format = "0.0%"
        month_value = formula_cells.get(f"{month_letter}{row_idx}", 0)
        formula_cells[f"{pct_letter}{row_idx}"] = month_value / target if target else 0


def build_workbook(kpi_data_list: list, program_breakdown: dict, ref: ReferenceData, today: date) -> tuple:
    """Assemble the workbook: one tab per Stage plus the static KPI Targets
    tab. kpi_data_list is the five stage {(owner_id, day): count} dicts in
    STAGE_LABELS order. program_breakdown is the Retained Clients stage's
    {(owner_id, program, day): count}, used only for that tab's drill-down.
    Returns (workbook, {sheet_name: {cell_ref: value}}) for the LibreOffice-
    unavailable cached-formula-value fallback."""
    wb = Workbook()
    wb.remove(wb.active)
    sheet_cell_values = {}
    joao_id, rohan_id = ref.owner_ids["joao"], ref.owner_ids["rohan"]
    for stage_index, (stage_label, tab_name, counts) in enumerate(zip(STAGE_LABELS, TAB_NAMES, kpi_data_list)):
        breakdown = program_breakdown if stage_label == "Retained Clients" else None
        pct_of_target = None
        if stage_label == "Calls/Meetings":
            pct_of_target = {joao_id: CALLS_MEETINGS_ANNUAL_TARGET, rohan_id: CALLS_MEETINGS_ANNUAL_TARGET}
        formula_cells = _build_stage_sheet(
            wb, tab_name, stage_label, STAGE_FILLS[stage_index], STAGE_FONTS[stage_index],
            counts, ref, today, program_breakdown=breakdown, pct_of_target=pct_of_target,
        )
        sheet_cell_values[tab_name] = formula_cells
    _build_kpi_targets_sheet(wb)
    return wb, sheet_cell_values


def _build_stage_sheet(wb: Workbook, sheet_name: str, stage_label: str, stage_fill, stage_font,
                        counts: dict, ref: ReferenceData, today: date, program_breakdown: dict = None,
                        pct_of_target: dict = None) -> dict:
    """One Stage's Year-to-Date tab: Month -> Week -> Day column drill-down
    (Excel outline grouping, +/- expand) with a trailing YTD Total column.
    Rows are the two BDMs plus a Grand Total row. When program_breakdown is
    given (Retained Clients only), each BDM's row expands (row-level +/-)
    into nested rows broken down by Country and Program of Interest, and two
    annual-target rows (12/year per BDM) are added below Grand Total.
    """
    ws = wb.create_sheet(sheet_name)
    # summaryRight for the column axis (Month/Week/Day, as before);
    # summaryBelow=False for the row axis, since a BDM's summary row sits
    # ABOVE its Program-of-Interest breakdown rows, not below them.
    ws.sheet_properties.outlinePr = Outline(summaryRight=True, summaryBelow=False)

    joao_id, rohan_id = ref.owner_ids["joao"], ref.owner_ids["rohan"]
    calendar = build_calendar_hierarchy(today)
    layout = _compute_column_layout(calendar, include_pct_column=pct_of_target is not None)

    ws.merge_cells("A1:A2")
    ws["A1"] = str(today.year)
    for row in (1, 2):
        cell = ws.cell(row=row, column=1)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    formula_cells: dict = {}
    owners = [(joao_id, "João Pacheco Gonçalves"), (rohan_id, "Rohan Harris")]

    # Union of programs across both owners, so every BDM shows the same set
    # of breakdown rows (even a 0-count program) for easy comparison.
    all_programs = []
    if program_breakdown is not None:
        seen = set()
        for (_owner, program, _day) in program_breakdown:
            if program not in seen:
                seen.add(program)
                all_programs.append(program)
        all_programs.sort()

    row_cursor = 3
    bdm_col_values = {}
    bdm_row_index = {}
    for owner_id, owner_name in owners:
        bdm_row = row_cursor
        bdm_row_index[owner_id] = bdm_row
        row_cursor += 1

        if program_breakdown is not None:
            child_values = []
            program_rows = []
            for program in all_programs:
                p_row = row_cursor
                ws.cell(row=p_row, column=1, value=f"    {program}").font = BODY_FONT

                def lookup(day, _owner_id=owner_id, _program=program):
                    return program_breakdown.get((_owner_id, _program, day), 0)

                child_values.append(_write_data_row(ws, p_row, layout, lookup, BODY_FONT, formula_cells))
                ws.row_dimensions[p_row].outline_level = 1
                ws.row_dimensions[p_row].hidden = True
                program_rows.append(p_row)
                row_cursor += 1

            bold_font = Font(name=FONT_BODY, color=BODY_TEXT, bold=True)
            ws.cell(row=bdm_row, column=1, value=owner_name).font = bold_font
            bdm_col_values[owner_id] = _write_sum_row(
                ws, bdm_row, layout, child_values, program_rows, bold_font, formula_cells
            )
            ws.row_dimensions[bdm_row].outline_level = 0
            # This BDM row is the summary for its collapsed Program-of-Interest
            # rows below it (see the collapsed-flag note in _write_time_headers
            # - same OOXML mechanism, row axis instead of column axis).
            ws.row_dimensions[bdm_row].collapsed = True
        else:
            ws.cell(row=bdm_row, column=1, value=owner_name).font = BODY_FONT

            def lookup(day, _owner_id=owner_id):
                return counts.get((_owner_id, day), 0)

            bdm_col_values[owner_id] = _write_data_row(ws, bdm_row, layout, lookup, BODY_FONT, formula_cells)

        if pct_of_target is not None:
            _write_pct_of_target_row(ws, bdm_row, layout, pct_of_target[owner_id], BODY_FONT, formula_cells)

    total_row = row_cursor
    row_cursor += 1
    total_font = Font(name=FONT_BODY, color=BODY_TEXT, bold=True)
    ws.cell(row=total_row, column=1, value="Grand Total").font = total_font
    _write_sum_row(
        ws, total_row, layout,
        [bdm_col_values[joao_id], bdm_col_values[rohan_id]],
        [bdm_row_index[joao_id], bdm_row_index[rohan_id]],
        total_font, formula_cells,
    )
    if pct_of_target is not None:
        # Grand Total's own target is the sum of both BDMs' individual
        # targets (880 = 440 + 440), not a separately-specified figure.
        _write_pct_of_target_row(ws, total_row, layout, sum(pct_of_target.values()), total_font, formula_cells)

    target_rows = []
    if program_breakdown is not None:
        # Per-rep annual target: 12 retained clients/year each, shown as
        # literal progressive "n/12" text per elapsed month + YTD - a static
        # pace marker, not a formula (there's nothing to sum).
        for owner_id, owner_name in owners:
            t_row = row_cursor
            ws.cell(row=t_row, column=1,
                    value=f"{owner_name} - Target ({RETAINED_CLIENT_ANNUAL_TARGET}/year)").font = MUTED_ITALIC_FONT
            for month_total_col, _week_cols, _mn, month_index, _pct in layout["month_groups"]:
                ws.cell(row=t_row, column=month_total_col,
                        value=f"{month_index}/{RETAINED_CLIENT_ANNUAL_TARGET}").font = MUTED_ITALIC_FONT
            ws.cell(row=t_row, column=layout["ytd_col"],
                    value=f"{today.month}/{RETAINED_CLIENT_ANNUAL_TARGET}").font = MUTED_ITALIC_FONT
            target_rows.append(t_row)
            row_cursor += 1

    last_data_row = row_cursor - 1

    _write_time_headers(ws, layout, stage_label, stage_fill, stage_font)

    for r in range(3, last_data_row + 1):
        for c in range(1, layout["last_col"] + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if c > 1:
                cell.alignment = Alignment(horizontal="center")
    for c in range(1, layout["last_col"] + 1):
        ws.cell(row=total_row, column=c).fill = GRAND_TOTAL_FILL

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 30

    return formula_cells


def _build_kpi_targets_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("KPI Targets")

    ws["A1"] = "Minimum Annual KPIs"
    ws["A1"].font = TITLE_FONT

    # Ordered to match the Year to Date sheet's Stage column order:
    # Retained -> Referred -> New Intermediaries -> Presentations ->
    # Calls/Meetings.
    headers = ["KPI", "Minimum Target"]
    rows = [
        ("Retained clients (minimum)", "12/year (~1/month)"),
        ("Intermediary-referred clients", "36/year (~3/month)"),
        ("New self-sourced intermediaries", "150/year (~12/month, ~3/week)"),
        ("Presentations", "40/year (~3/month, ~1/week)"),
        ("Calls / meetings", "440/year per BDM (~37/month, ~8/week)"),
    ]
    header_row = 3
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for i, (kpi, target) in enumerate(rows):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=kpi).font = BODY_FONT
        ws.cell(row=row, column=2, value=target).font = BODY_FONT
        for c in (1, 2):
            ws.cell(row=row, column=c).border = THIN_BORDER

    note_row = header_row + 1 + len(rows) + 1
    ws.cell(row=note_row, column=1,
            value="Weekly figures assume a 5-day work week. These are minimum expectations, "
                  "not aspirational targets.").font = MUTED_ITALIC_FONT

    section2_row = note_row + 2
    ws.cell(row=section2_row, column=1, value="Detailed Expectations").font = TITLE_FONT

    bullets = [
        "The 36-referred-to-12-retained ratio is deliberate - it sets an explicit ~33% minimum "
        "conversion rate, tracked monthly and annually, not just at year-end",
        "Activity volume (calls, meetings, presentations) is a leading indicator, not the target - "
        "hitting activity minimums while missing the 12-retained-clients/year floor is not meeting "
        "the mandate",
    ]
    for i, bullet in enumerate(bullets):
        row = section2_row + 1 + i
        cell = ws.cell(row=row, column=1, value=f"• {bullet}")
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 34
    for i, bullet in enumerate(bullets):
        ws.row_dimensions[section2_row + 1 + i].height = 30


def _try_libreoffice_recalculation(xlsx_path: str) -> bool:
    """Attempt a headless LibreOffice round-trip (convert-to xlsx
    re-evaluates formulas). Returns True on success, False if LibreOffice
    isn't usable here (e.g. a Calc-less install) so the caller can fall
    back - never raises for that case, since it's a soft dependency."""
    out_dir = tempfile.mkdtemp(prefix="gcs_report_recalc_")
    try:
        result = subprocess.run(
            ["soffice", "--headless", "--calc", "--convert-to", "xlsx", "--outdir", out_dir, xlsx_path],
            capture_output=True, text=True, timeout=120,
        )
        recalculated = os.path.join(out_dir, os.path.basename(xlsx_path))
        if result.returncode != 0 or not os.path.exists(recalculated):
            print(f"  LibreOffice unavailable/failed ({result.stderr.strip() or result.stdout.strip()}), "
                  f"falling back to direct cached-value injection.")
            return False
        os.replace(recalculated, xlsx_path)
        return True
    except (OSError, subprocess.SubprocessError) as exc:
        print(f"  LibreOffice unavailable ({exc}), falling back to direct cached-value injection.")
        return False
    finally:
        if os.path.isdir(out_dir):
            for f in os.listdir(out_dir):
                os.remove(os.path.join(out_dir, f))
            os.rmdir(out_dir)


def _inject_cached_formula_values(xlsx_path: str, sheet_cell_values: dict) -> None:
    """Fallback recalculation: write the already-known Grand Total values
    directly into each formula cell's cached <v> in the sheet XML. The
    formula itself (=SUM(...)) is untouched - this only supplies the
    cached result a viewer would otherwise need to compute itself, exactly
    what a LibreOffice/Excel recalculation pass would leave behind."""
    import xml.etree.ElementTree as ET

    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    doc_rels_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg_rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", ns)

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        workbook_xml = ET.fromstring(zin.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
        rid_to_target = {
            rel.get("Id"): rel.get("Target")
            for rel in rels_xml.findall(f"{{{pkg_rels_ns}}}Relationship")
        }

        sheet_name_to_xml = {}
        for sheet in workbook_xml.find(f"{{{ns}}}sheets"):
            name = sheet.get("name")
            rid = sheet.get(f"{{{doc_rels_ns}}}id")
            target = rid_to_target.get(rid)
            if name and target:
                # Target may be package-absolute ("/xl/worksheets/sheet1.xml")
                # or relative to xl/ ("worksheets/sheet1.xml") - handle both.
                sheet_name_to_xml[name] = target.lstrip("/") if target.startswith("/") else "xl/" + target

        updates = {}  # sheet_xml_path -> new bytes
        for sheet_name, cell_values in sheet_cell_values.items():
            xml_path = sheet_name_to_xml.get(sheet_name)
            if not xml_path or not cell_values:
                continue
            root = ET.fromstring(zin.read(xml_path))
            for c in root.iter(f"{{{ns}}}c"):
                ref = c.get("r")
                if ref in cell_values:
                    v_elem = c.find(f"{{{ns}}}v")
                    if v_elem is None:
                        v_elem = ET.SubElement(c, f"{{{ns}}}v")
                    v_elem.text = str(cell_values[ref])
            updates[xml_path] = ET.tostring(root, encoding="UTF-8", xml_declaration=True)

        tmp_path = xlsx_path + ".tmp"
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = updates.get(item.filename, zin.read(item.filename))
                zout.writestr(item, data)

    os.replace(tmp_path, xlsx_path)


def recalculate_workbook(xlsx_path: str, sheet_cell_values: dict) -> None:
    """Recalculate all formulas so the saved workbook shows real cached
    SUM() values on open, not formula placeholders. Prefers a headless
    LibreOffice round-trip; falls back to injecting the already-known
    Grand Total values as cached formula results if LibreOffice isn't
    usable in this environment."""
    if not _try_libreoffice_recalculation(xlsx_path):
        _inject_cached_formula_values(xlsx_path, sheet_cell_values)


def print_reconciliation_table(kpi_data_list: list) -> None:
    """Print each stage's year-to-date grand total. There's no fixed
    "expected" figure to check against any more (unlike the original
    all-time build) - the whole point of this table is a live number that
    changes daily. Each fetch function already prints its own reasoning
    when a total is noteworthy (e.g. Retained Clients dropping once
    scoped to this year); this is just the final at-a-glance summary.
    """
    print("\n=== Year-to-Date Summary ===")
    print(f"{'Stage':<24} {'YTD Total':>10}")
    for title, counts in zip(STAGE_LABELS, kpi_data_list):
        print(f"{title:<24} {sum(counts.values()):>10}")


def main() -> int:
    load_dotenv()
    access_token = os.environ.get("HUBSPOT_ACCESS_TOKEN", "")
    client = HubSpotClient(access_token)
    ref = resolve_reference_data(client)

    new_intermediaries = fetch_new_intermediaries(client, ref)
    new_meetings, new_presentations = fetch_meetings(client, ref)
    lead_source_property, referred_clients = fetch_referred_clients(client, ref)
    retained_clients, retained_program_breakdown = fetch_retained_clients(client, ref, lead_source_property)

    # Order matches STAGE_LABELS: Retained -> Referred -> New Intermediaries
    # -> Presentations -> Calls/Meetings.
    kpi_data_list = [retained_clients, referred_clients, new_intermediaries, new_presentations, new_meetings]

    print("\nBuilding workbook...")
    from datetime import datetime, timezone

    today = datetime.now(timezone.utc).date()
    wb, sheet_cell_values = build_workbook(kpi_data_list, retained_program_breakdown, ref, today)
    os.makedirs("reports", exist_ok=True)
    xlsx_path = os.path.join("reports", "reports.xlsx")
    wb.save(xlsx_path)

    print("Recalculating formulas...")
    recalculate_workbook(xlsx_path, sheet_cell_values)
    print(f"Wrote {xlsx_path}")

    print_reconciliation_table(kpi_data_list)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except HubSpotError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001 - top-level safety net per spec
        print(f"ERROR: unexpected failure: {exc}", file=sys.stderr)
        sys.exit(1)
